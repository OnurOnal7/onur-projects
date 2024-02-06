from tkinter import messagebox
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# Time-Entry Automation for iLab.
#
# Author: Onur Onal


# Calls the time entry automation.
def run_time_entry(excel_file_path):
    # Retrieves the range of Excel data from a  filepath.
    def get_excel_range(file_path):
        # Pulls Excel workbook and worksheet from file path.
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook['Sheet1']
        return worksheet.iter_rows(min_row=5, max_row=worksheet.max_row, min_col=1, max_col=11)

    # Retrieves column values of a given row.
    def get_column_values(row):
        return [cell.value for cell in row]

    # Navigates to the given url.
    def navigate_to_url(driver, url):
        driver.get(url)

    # Finds an elements by its XPATH.
    def find_by_xpath(source, xpath):
        return source.find_element(By.XPATH, xpath)

    # Finds an elements by its CSS Selector.
    def find_by_css(source, css):
        return source.find_element(By.CSS_SELECTOR, css)

    # Creates a wait until an element becomes clickable.
    def wait_until_clickable(source, t, xpath):
        return WebDriverWait(source, t).until(EC.element_to_be_clickable((By.XPATH, xpath)))

    # Clicks the sign-in button and waits for the user to enter login credentials.
    def sign_in(driver):
        link = find_by_xpath(driver, "//*[@id='integration_login']/div[1]/p/a")
        link.click()
        wait_until_clickable(driver, 3600, "//a[normalize-space()='Iowa State University']")

    # Selects the current staff member from the dropdown and clicks the "Add New Row" button.
    def add_new_row(driver, row):
        # Waits for the time entry page to load and stores the initial selected staff member in the dropdown.
        button = wait_until_clickable(driver, 10, "//*[@id='timesheets-table']/thead[1]/tr[1]/th[2]/div/button")
        selected = find_by_xpath(driver, "//div[@class='text']").text

        # Selects the staff member from the dropdown if not initially selected.
        if selected != row[0]:
            dropdown = wait_until_clickable(driver, 10, "//*[@id='coreAssociateSelect']/input")
            dropdown.click()
            time.sleep(0.1)
            option = wait_until_clickable(driver, 10, f"//div[normalize-space()='{row[0]}']")
            option.click()
            time.sleep(1.5)

        # Clicks the "Add New Row" button.
        time.sleep(1)
        button.click()

    # Enters each job into the "Job" text box.
    def enter_jobs(driver, jobs):
        for job in jobs:
            text = find_by_css(driver, "ul.select2-choices input.select2-input")
            text.send_keys(job)
            option = wait_until_clickable(driver, 10, f"//div[contains(@class, 'result')]/span[text()='{job}']/../..")
            option.click()

    # Enters each service into the "Service" text box.
    def enter_services(driver, services):
        # Unchecks "Select all".
        checkbox = find_by_xpath(driver, ".//*[contains(text(), 'Select all')]")
        checkbox.click()
        dropdown = find_by_xpath(driver, "//div[@role='combobox']//i[@class='dropdown icon']")
        dropdown.click()

        # Finds and selects each service from the "Service" dropdown.
        for service in services:
            option = find_by_xpath(driver, f"//span[normalize-space()='{service}']")
            option.click()

        dropdown.click()

    # Finds and clicks on the "Add" button.
    def save_new_row(driver):
        button = find_by_xpath(driver, "//button[@class='ui primary button']")
        button.click()

    # Returns a dictionary consisting of each job-service pair in the correct ordering.
    def populate_entries(jobs, services):
        # Dictionary that stores each job-service pair in the correct order.
        # Key: Job-Service concatenation.
        # Value: Order of the job-service pair.
        entries = {}
        i = 0

        # Populates the dictionary, key being the job-service pair and value being its order.
        for job in jobs:
            j = 0
            for service in services:
                key = f"{job}, {service}"  # Concatenates the job and services.
                value = i * len(services) + j  # Calculates the order of the pair being processed.
                entries[key] = value  # Creates the key-value pair.
                j += 1
            i += 1

        return entries

    # Returns a 2D list containing of all the time entries made for each day.
    def populate_days(row):
        days = []  # 2D list that stores each entry for each day in the worksheet.

        # Populates the 2D list by splitting the values for each day.
        for i in range(4, 11):
            days.append(row[i].split(", "))

        return days

    # Navigates to the correct week which the time entry will be made.
    def find_week(driver, row):
        wait = 2
        k = 0

        # Clicks the next week button until the desired week is reached.
        while k < row[3]:
            index = 3 if k == 0 else 4
            button = wait_until_clickable(driver, 10, f"//*[@id='timesheets-table']/thead[1]/tr[1]/th[3]/i[{index}]")
            time.sleep(1)
            button.click()
            k += 1
            wait = 3

        time.sleep(wait)

    # Updates the outer key of the "lines" dictionary.
    def update_lines(driver, row, lines, entries):
        i = 1
        j = max(entries.values())  # Set to the number of job-service pairs.

        # Sets the current staff member as an outer key in the "lines" dictionary.
        if row[0] not in lines:
            lines[row[0]] = {}

        # Updates the key-value pairs for the outer key of the "lines dictionary".
        # For staff members with previously created rows, this extracts the correct line for the rows that have most recently been created.
        while j > -1:
            # Sequentially retrieves lines from the table of rows.
            line = driver.find_element(By.XPATH, f"//*[@id='timesheets-table']/tbody[{i}]/tr[1]")
            i += 1

            # If the line already exists, its key-value pair is removed.
            # If the line does not exist, it becomes the inner key, with a value of its job-service concatenation.
            if line in lines[row[0]]:
                lines[row[0]].pop(line)
            else:
                current_job = line.find_element(By.XPATH, ".//td[1]/a").text
                current_service = line.find_element(By.XPATH, ".//td[3]").text
                current_entry = f"{current_job}, {current_service}"
                lines[row[0]][line] = current_entry
                j -= 1

    # Carries out time entry for each row in the table.
    def time_entry(lines, row, entries, days):
        for line, current in lines[row[0]].items():
            boxes = []  # The seven boxes that are present in the current row.

            # The box corresponding to each day of the week is found and added to the list.
            for i in range(4, 11):
                box = find_by_xpath(line, f".//td[{i}]/div/span/b[1]")
                boxes.append(box)

            # Finds the appropriate row in the table.
            for entry, index in entries.items():
                # Searches for the row by its job-service concatenation.
                if entry == current:
                    # Adds time for each day in the current row.
                    for box, day in zip(boxes, days):
                        box_val = day[index]  # Retrieves a value from each day's list.
                        clicks = int(float(box_val) / 0.25)  # Calculates the amount of arrow clicks needed to reach the value.

                        # Enacts the clicks for each box in the row.
                        for _ in range(clicks):
                            box.click()
                    break  # Loop is terminated when correct row is found.

    # Save time-entry and navigates back to the current week.
    def save_time_entry(driver, row):
        time.sleep(2)
        button = find_by_xpath(driver, "//*[@id='timesheets-action-bar']/a[2]")
        button.click()
        wait = 6

        # Returns to the current week if time was entered in a future week.
        if row[3] > 0:
            time.sleep(6)
            button = wait_until_clickable(driver, 10, "//*[@id='timesheets-table']/thead[1]/tr[1]/th[3]/i[2]")
            button.click()
            wait = 2

        time.sleep(wait)


    def main():
        data = get_excel_range(excel_file_path)
        rows = []

        # Populates the list with the column values of each row.
        for row in data:
            column_value = get_column_values(row)
            rows.append(column_value)

        print(len(rows))

        # Instantiates a web driver object and navigates to the URL of the time entry page in iLab.
        driver = webdriver.Chrome()
        driver.maximize_window()  # Maximizes the browser window.
        navigate_to_url(driver, "https://iastate.ilab.agilent.com/service_centers/5171/timesheets")

        try:
            sign_in(driver)

            # Nested dictionary that will store the rows for each staff member.
            # Outer key: Staff member name.
            # Inner key: Time entry row.
            # Value: Job-Service concatenation.
            lines = {}

            # Repeats time-entry process for each row of the Excel worksheet.
            for row in rows:
                add_new_row(driver, row)

                jobs = row[1].split(", ")  # Stores the jobs in a list.
                enter_jobs(driver, jobs)  # Enters the jobs to the "Job" text box.

                services = row[2].split(", ")  # Stores the services in a list.
                enter_services(driver, services)  # Enters the services to the "Service" text box.

                save_new_row(driver)

                entries = populate_entries(jobs, services)  # Dictionary of all the current job-service pairs with their correct order.

                days = populate_days(row)  # 2D list of all the time entries for each day.

                find_week(driver, row)  # Navigates to the appropriate week.

                update_lines(driver, row, lines, entries)  # Updates the "lines" dictionary.

                time_entry(lines, row, entries, days)  # Carries out the time entry for each row in the table.

                save_time_entry(driver, row)  # Saves the time entry and resets.

        except Exception as e:
            print("An error occurred:", e)

        finally:
            # Displays completion message, closes browser window, and ends WebDriver Session.
            messagebox.showinfo("Completion Message", "All time entries have been completed.")
            driver.quit()

    main()
